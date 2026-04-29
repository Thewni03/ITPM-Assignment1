import openpyxl
import os

EXCEL_PATH = os.path.expanduser("~/Downloads/test_automation/Assignment 1 - Test cases.xlsx")

REPLACEMENTS = {
    2:   ("Neg_0001", "M", "oyage birthday eka kawadada kiyala dannawada?", "ඔයාගේ birthday එක කවදාද කියලා දන්නවද?"),
    22:  ("Neg_0006", "M", "machan bohoma kaalakin! kohomada oya dhan?", "මචන් බොහොම කාලකින්! කොහොමද ඔයා දැන්?"),
    26:  ("Neg_0007", "M", "karunakarala mage bag eka poddak balahinna.", "කරුණාකරලා මගේ bag එක පොඩ්ඩක් බලාහින්න."),
    38:  ("Neg_0010", "M", "aiyo, eka wage wunoth amaru thamai machan.", "අයියෝ, ඒක වාගේ උනොත් අමාරු තමයි මචන්."),
    42:  ("Neg_0011", "S", "yana yana thenata phone karanna epa.", "යන යන තැනට phone කරන්න එපා."),
    46:  ("Neg_0012", "M", "hamu hamu wenakota katha karamu api.", "හමු හමු වෙනකොට කතා කරමු අපි."),
    50:  ("Neg_0013", "M", "mama yanawa... oyath enna? api ekka yanawa!", "මම යනවා... ඔයාත් එන්නද? අපි එක්ක යනවා!"),
    78:  ("Neg_0020", "M", "keep in touch machan, missing you all so much.", "keep in touch මචන්, missing you all so much."),
    86:  ("Neg_0022", "M", "oyage laptop eke bluetooth on karala file eka share karanna.", "ඔයාගේ laptop එකේ bluetooth on කරලා file එක share කරන්න."),
    90:  ("Neg_0023", "M", "YouTube eke subscribe karanna, TikTok eke follow karanna.", "YouTube එකේ subscribe කරන්න, TikTok එකේ follow කරන්න."),
    106: ("Neg_0027", "M", "ape prof kiwa heta lab eka cancel kiyala.", "අපේ prof කිව්වා හෙට lab එක cancel කියලා."),
    114: ("Neg_0029", "M", "Matara walata yanawa, Mirissa beach ekatath yamuda?", "මාතර වලට යනවා, Mirissa beach එකටත් යමුද?"),
    118: ("Neg_0030", "M", "Nugegoda junction walata enna, Maharagama passe.", "නුගේගොඩ junction වලට එන්න, මහරගම පස්සේ."),
    134: ("Neg_0034", "M", "ape 1st year batch eke 200k withara lamai innawa.", "අපේ 1st year batch එකේ 200ක් විතර ළමයි ඉන්නවා."),
    142: ("Neg_0036", "M", "phone eka GBP 299 walin gaththa duty free eken.", "phone එක GBP 299 වලින් ගත්තා duty free එකෙන්."),
    150: ("Neg_0038", "M", "heta 6:00AM walata nathanawa, 11:30PM walata nidagannam.", "හෙට 6:00AM වලට නඟිනවා, 11:30PM වලට නිදාගන්නම්."),
    154: ("Neg_0039", "M", "event eka May 5 walata. June 10 wenakan register karanna.", "event එක May 5 වලට. June 10 වෙනකන් register කරන්න."),
    170: ("Neg_0043", "M", "ado meka godak mara scene ekak. set thamai bro.", "අඩෝ මේක ගොඩක් මාර scene එකක්. සෙට් තමයි bro."),
    178: ("Neg_0045", "M", "nawa blog eka balanna: www.islandlife.lk/travel kiyala.", "නව blog එක බලන්න: www.islandlife.lk/travel කියලා."),
    182: ("Neg_0046", "M", "@Tharindu ayya heta meeting eken passe online inna puluwan da?", "@තාරිඳු අයියා හෙට meeting එකෙන් පස්සේ online ඉන්න පුළුවන්ද?"),
    194: ("Neg_0049", "M", "bro heta 7:45AM walata lab, 12:30PM exam. exhausted already 😩", "bro හෙට 7:45AM වලට lab, 12:30PM exam. exhausted already 😩"),
}

def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # columns are fixed based on inspection
    COL_TCID     = 1
    COL_LENGTH   = 2
    COL_INPUT    = 3
    COL_EXPECTED = 4
    COL_ACTUAL   = 5
    COL_STATUS   = 6

    replaced = 0
    for row_num, (tc_id, length, inp, expected) in REPLACEMENTS.items():
        ws.cell(row=row_num, column=COL_TCID).value     = tc_id
        ws.cell(row=row_num, column=COL_LENGTH).value   = length
        ws.cell(row=row_num, column=COL_INPUT).value    = inp
        ws.cell(row=row_num, column=COL_EXPECTED).value = expected
        ws.cell(row=row_num, column=COL_ACTUAL).value   = None
        ws.cell(row=row_num, column=COL_STATUS).value   = None
        print(f"  Row {row_num} replaced -> {tc_id}: {inp[:50]}")
        replaced += 1

    wb.save(EXCEL_PATH)
    print(f"\nDone! Replaced {replaced} rows and saved.")

if __name__ == "__main__":
    main()
