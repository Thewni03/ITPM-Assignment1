import openpyxl
import os

EXCEL_PATH = os.path.expanduser("~/Downloads/test_automation/Assignment 1 - Test cases.xlsx")

DATA = {
    2:  ("Question forms",
         "Rationale: The input is phrased as a question asking about someone's birthday date."),
    6:  ("Question forms",
         "Rationale: The input contains two questions — asking if someone arrived on time and how the trip was."),
    10: ("Command forms",
         "Rationale: The input gives a direct command to leave quickly or face being late."),
    14: ("Command forms",
         "Rationale: The input instructs someone to do something early, otherwise the speaker will do it."),
    18: ("Greetings",
         "Rationale: The input is a New Year greeting wishing someone well."),
    22: ("Greetings",
         "Rationale: The input is a casual greeting after a long time, asking how the person is."),
    26: ("Requests",
         "Rationale: The input politely requests someone to watch over a bag."),
    30: ("Requests",
         "Evidence: 'meka poddak balanna puluwanda?' — this is a polite request asking someone to check something."),
    34: ("Responses",
         "Rationale: The input is a negative response denying knowledge of something."),
    38: ("Responses",
         "Rationale: The input is a casual response expressing that something would be difficult if it were true."),
    42: ("Repeated Words",
         "Evidence: 'hari hari' — a repeated word used for emphasis in casual Singlish."),
    46: ("Repeated Words",
         "Evidence: 'denna denna' — a repeated phrase used to express regularity or frequency."),
    50: ("Inputs with Punctuation Marks",
         "Rationale: The input uses '...' for trailing thought and '?' and '!' as punctuation marks."),
    54: ("Inputs with Punctuation Marks",
         "Rationale: The input uses '...' for pause effect and '?' as a question mark."),
    58: ("Romanization / Spelling Variants",
         "Evidence: 'machng' is a shortened spelling variant of 'machan' and 'ennam' is a variant of 'ennam'."),
    62: ("Romanization / Spelling Variants",
         "Evidence: 'pulwn' is a shortened variant of 'puluwana' and 'kraanna' is a variant of 'karanna'."),
    66: ("Isolated English Word Insertions in Singlish",
         "Evidence: 'lunch' and 'hotel' are single English words inserted within a Singlish sentence."),
    70: ("Isolated English Word Insertions in Singlish",
         "Evidence: 'release', 'series', and 'plot' are single English words inserted within a Singlish sentence."),
    74: ("Multi-Word English Phrases in Singlish",
         "Evidence: 'out of office' is a multi-word English phrase inserted within the Singlish sentence."),
    78: ("Multi-Word English Phrases in Singlish",
         "Evidence: 'lets catch up this weekend' is a multi-word English phrase used within the Singlish sentence."),
    82: ("English Digital Terms in Singlish",
         "Evidence: 'RAM' is a digital/technical term inserted within the Singlish sentence."),
    86: ("English Digital Terms in Singlish",
         "Evidence: 'screenshare' is a digital term inserted within the Singlish sentence."),
    90: ("Platform/App Names in Singlish",
         "Evidence: 'YouTube' and 'TikTok' are platform names inserted within the Singlish sentence."),
    94: ("Platform/App Names in Singlish",
         "Evidence: 'Instagram' and 'Facebook' are platform names inserted within the Singlish sentence."),
    98: ("English Abbreviations/Acronyms in Singlish",
         "Evidence: 'DOB' is an abbreviation for Date of Birth inserted within the Singlish sentence."),
    102: ("English Abbreviations/Acronyms in Singlish",
          "Evidence: 'CV', 'PDF', and 'ASAP' are abbreviations/acronyms inserted within the Singlish sentence."),
    106: ("English Clipped Forms in Singlish",
          "Evidence: 'prof' is a clipped form of 'professor' and 'lab' is a clipped form of 'laboratory'."),
    110: ("English Clipped Forms in Singlish",
          "Evidence: 'prof' is a clipped form of 'professor', 'sub' is a clipped form of 'submit', and 'deadline' is used as a clipped English term."),
    114: ("Place Names Embedded in Singlish",
          "Evidence: 'Matara' and 'Mirissa' are Sri Lankan place names embedded within the Singlish sentence."),
    118: ("Place Names Embedded in Singlish",
          "Evidence: 'Nugegoda' and 'Maharagama' are Sri Lankan place names embedded within the Singlish sentence."),
    122: ("Person Names Embedded in Singlish",
          "Evidence: 'Kasun' is a Sri Lankan person name embedded within the Singlish sentence."),
    126: ("Person Names Embedded in Singlish",
          "Evidence: 'Nimasha' and 'Dineth' are Sri Lankan person names embedded within the Singlish sentence."),
    130: ("Inputs with Numbers and Numeric Suffixes",
          "Evidence: '3rd' is a numeric suffix used within the Singlish sentence."),
    134: ("Inputs with Numbers and Numeric Suffixes",
          "Evidence: '1st', '2nd', '3rd', '4th' are numeric suffixes and '300k' and '150k' are numbers used within the Singlish sentence."),
    138: ("Inputs with Currency",
          "Evidence: 'Rs. 500' is a currency value embedded within the Singlish sentence."),
    142: ("Inputs with Currency",
          "Evidence: 'GBP 299' is a foreign currency value embedded within the Singlish sentence."),
    146: ("Inputs with Time Formats",
          "Evidence: '9:00AM' is a time format embedded within the Singlish sentence."),
    150: ("Inputs with Time Formats",
          "Evidence: '6:00AM' and '11:30PM' are time formats embedded within the Singlish sentence."),
    154: ("Inputs with Dates",
          "Evidence: 'May 5' and 'June 10' are date references embedded within the Singlish sentence."),
    158: ("Inputs with Dates",
          "Evidence: 'March 31' and 'April 1' are date references embedded within the Singlish sentence."),
    162: ("Inputs with Unit of Measurements",
          "Evidence: 'feet 20k' is a unit of measurement embedded within the Singlish sentence."),
    166: ("Inputs with Unit of Measurements",
          "Evidence: 'km 5k' and 'minutes' are units of measurement embedded within the Singlish sentence."),
    170: ("Inputs with Slang and Casual Phrasing",
          "Evidence: 'mara scene' and 'set thamai' are slang expressions used in casual Singlish."),
    174: ("Inputs with Slang and Casual Phrasing",
          "Evidence: 'godak niyamai' and 'hodata hadala' are casual Singlish slang expressions."),
    178: ("Online Identifiers in Singlish",
          "Evidence: 'www.islandlife.lk/travel' is a website URL embedded within the Singlish sentence."),
    182: ("Online Identifiers in Singlish",
          "Evidence: '@Tharindu' is a social media handle/mention embedded within the Singlish sentence."),
    186: ("Inputs Containing Emojis",
          "Evidence: '😂' is an emoji used at the end of the Singlish sentence to express laughter."),
    190: ("Inputs Containing Emojis",
          "Evidence: '😤' is an emoji used mid-sentence to express frustration in the Singlish input."),
    194: ("Inputs with Time Formats, English Clipped Forms in Singlish, Inputs Containing Emojis",
          "Evidence: '7:45AM' and '12:30PM' are time formats. 'lab' is a clipped form of 'laboratory'. '😩' is an emoji expressing exhaustion."),
    198: ("Inputs with Time Formats, Inputs with Currency, Person Names Embedded in Singlish, English Abbreviations/Acronyms in Singlish, Inputs Containing Emojis",
          "Evidence: '9:00AM' is a time format. 'Rs.1500' is a currency value. 'Sachini' is a person name. 'ASAP' is an abbreviation. '🤞' is an emoji expressing hope."),
}

def main():
    print(f"Opening: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.cell(row=1, column=7).value = "Singlish input types covered"
    ws.cell(row=1, column=8).value = "Evidence or rationale for the input type covered"

    filled = 0
    for row_num, (types, rationale) in DATA.items():
        ws.cell(row=row_num, column=7).value = types
        ws.cell(row=row_num, column=8).value = rationale
        print(f"  Row {row_num} -> {types[:50]}")
        filled += 1

    wb.save(EXCEL_PATH)
    print(f"\nDone! Filled {filled} rows with types and rationale.")

if __name__ == "__main__":
    main()
